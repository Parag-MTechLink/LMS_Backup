"""
RFQ upload route: Excel → canonical JSON → schema validation → business validation → workflow.
Orchestration only; no business logic. File type/size validated at boundary.
"""
import io
import os
import tempfile
from typing import List

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel
from openpyxl import Workbook

from app.core.config import settings
from app.core.constants import RFQ_ALLOWED_EXTENSIONS
from app.core.database import SessionLocal
from app.core.logging_config import (
    log_rfq_upload_error,
    log_rfq_upload_start,
    log_rfq_upload_success,
    log_rfq_upload_validation_failed,
)
from app.services.excel_parser import parse_rfq_excel
from app.services.rfq_schema import validate_canonical_rfq
from app.services.rfq_upload_sync import sync_rfq_to_list
from app.services.rfq_validator import validate_business_rules
from app.services.rfq_workflow import execute_workflow

router = APIRouter(prefix="/rfq", tags=["RFQ Upload"])

# Template layout must match app.services.excel_parser (row 2 = header fields, from row 6 = test items)
RFQ_TEMPLATE_HEADER_ROW = {
    "A1": "Company Name", "B1": "Contact Person", "C1": "Email", "D1": "Phone",
    "E1": "Project Title", "F1": "Testing Type", "G1": "Deadline (YYYY-MM-DD)", "H1": "Urgent (Yes/No)",
}
RFQ_TEMPLATE_EXAMPLE_ROW2 = {
    "A2": "Acme Labs", "B2": "Jane Doe", "C2": "jane@acme.com", "D2": "+91 9876543210",
    "E2": "Safety Testing - Batch 2024", "F2": "Environmental", "G2": "2025-03-15", "H2": "No",
}
RFQ_TEMPLATE_TABLE_HEADERS = {"A5": "Test Name", "B5": "Standard", "C5": "Quantity"}
RFQ_TEMPLATE_TABLE_EXAMPLE = {"A6": "Temperature Test", "B6": "IEC 60068", "C6": 10}


def _build_rfq_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    if ws:
        ws.title = "RFQ"
        for cell, value in RFQ_TEMPLATE_HEADER_ROW.items():
            ws[cell] = value
        for cell, value in RFQ_TEMPLATE_EXAMPLE_ROW2.items():
            ws[cell] = value
        for cell, value in RFQ_TEMPLATE_TABLE_HEADERS.items():
            ws[cell] = value
        for cell, value in RFQ_TEMPLATE_TABLE_EXAMPLE.items():
            ws[cell] = value
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.get("/template")
def download_rfq_template():
    """Download Excel template for RFQ upload. Row 2 = company/contact/project/deadline; from row 6 = test name, standard, quantity."""
    content = _build_rfq_template_xlsx()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rfq_upload_template.xlsx"},
    )


class UploadIncompleteResponse(BaseModel):
    status: str = "Incomplete"
    errors: List[str]
    message: str = "Validation failed. Please fix errors and re-upload or edit."


class UploadSuccessResponse(BaseModel):
    status: str = "Pending Review"
    message: str = "RFQ submitted successfully."
    request_id: str | None = None
    rfq_id: int | None = None  # ID in main rfqs list (for display/navigation)


@router.post("/upload")
async def upload_rfq_excel(file: UploadFile = File(...)):
    """
    1. Validate file type and size
    2. Parse to canonical JSON
    3. JSON Schema validation
    4. Business rule validation
    5. Workflow (insert only after validation)
    6. Return Incomplete with errors or Pending Review
    """
    filename = file.filename or ""
    if not filename or not any(filename.lower().endswith(ext) for ext in RFQ_ALLOWED_EXTENSIONS):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted.")

    content = await file.read()
    file_size = len(content)
    if file_size > settings.MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File size exceeds maximum allowed ({settings.MAX_FILE_SIZE} bytes).",
        )

    log_rfq_upload_start(filename, file_size)

    path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(content)
            path = tmp.name

        canonical = parse_rfq_excel(path)
    except Exception as e:
        log_rfq_upload_error("parse", str(e))
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {str(e)}")
    finally:
        if path and os.path.isfile(path):
            try:
                os.unlink(path)
            except OSError:
                pass

    schema_valid, schema_errors = validate_canonical_rfq(canonical)
    business_result = validate_business_rules(canonical)

    db = SessionLocal()
    try:
        status, errors, rfq = execute_workflow(db, canonical, schema_errors, business_result)
        if status == "Incomplete":
            log_rfq_upload_validation_failed(filename, len(errors))
            return UploadIncompleteResponse(
                status="Incomplete",
                errors=errors,
                message="Validation failed. Please fix errors and re-upload or edit.",
            )
        # Sync to main rfqs list so it appears on RFQs page (find/create Customer, create RFQ)
        rfq_row = sync_rfq_to_list(db, canonical, rfq)
        if rfq_row:
            db.commit()
        log_rfq_upload_success(rfq.id, filename)
        return UploadSuccessResponse(
            status="Pending Review",
            message="RFQ submitted successfully.",
            request_id=str(rfq.id),
            rfq_id=rfq_row.id if rfq_row else None,
        )
    except Exception as e:
        db.rollback()
        log_rfq_upload_error("workflow", str(e))
        raise HTTPException(status_code=500, detail="Failed to save RFQ request.")
    finally:
        db.close()
