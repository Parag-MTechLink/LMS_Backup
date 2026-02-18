"""
Deterministic Excel parser for RFQ uploads.
Extraction only: produces canonical JSON. No DB, no validation, no status.
"""
import logging
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Fixed template: first sheet, row 2 = header field values
CELL_MAP = {
    "company_name": "A2",
    "contact_person": "B2",
    "email": "C2",
    "phone": "D2",
    "project_title": "E2",
    "testing_type": "F2",
    "deadline": "G2",
}
# Optional: urgent flag (e.g. H2 = "Yes" / "True")
URGENT_CELL = "H2"
# Test items table: data from row 6
TABLE_START_ROW = 6
MAX_TABLE_ROWS = 200


def _cell_value(ws, cell_ref: str) -> str:
    """Get string value from cell; empty if missing."""
    try:
        val = ws[cell_ref].value
        if val is None:
            return ""
        return str(val).strip()
    except Exception:
        return ""


def _parse_date(s: str) -> str | None:
    """Return ISO date string (YYYY-MM-DD) or None."""
    if not s or not s.strip():
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date().isoformat()
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.date().isoformat()
    except Exception:
        pass
    return None


def _parse_quantity(s: str) -> float:
    """Parse quantity; return 0 if invalid."""
    if not s or not str(s).strip():
        return 0.0
    try:
        return float(str(s).strip().replace(",", "."))
    except ValueError:
        return 0.0


def parse_rfq_excel(file_path: str) -> dict[str, Any]:
    """
    Read Excel and return canonical JSON only.
    Responsibility: extraction only. No DB, no validation, no status.
    """
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        wb.close()
        raise ValueError("Workbook has no active sheet")

    out: dict[str, Any] = {
        "company_name": "",
        "contact_person": "",
        "email": "",
        "phone": "",
        "project_title": "",
        "testing_type": "",
        "sample_quantity": 0,
        "deadline": None,
        "urgent_flag": False,
        "test_items": [],
    }

    for field, cell_ref in CELL_MAP.items():
        val = _cell_value(ws, cell_ref)
        out[field] = val
        if field == "deadline" and val:
            out["deadline"] = _parse_date(val)

    urgent_val = _cell_value(ws, URGENT_CELL).lower()
    out["urgent_flag"] = urgent_val in ("yes", "true", "1", "y")

    total_qty = 0.0
    for row_idx in range(TABLE_START_ROW, min(TABLE_START_ROW + MAX_TABLE_ROWS, ws.max_row + 1)):
        test_name = _cell_value(ws, f"A{row_idx}")
        standard = _cell_value(ws, f"B{row_idx}")
        qty_str = _cell_value(ws, f"C{row_idx}")
        quantity = _parse_quantity(qty_str)
        if not test_name and not standard and quantity == 0:
            continue
        out["test_items"].append({
            "test_name": test_name,
            "standard": standard,
            "quantity": quantity,
        })
        total_qty += quantity

    out["sample_quantity"] = int(total_qty) if total_qty == int(total_qty) else int(round(total_qty))
    wb.close()
    logger.info("Parsed canonical RFQ JSON: %d test items, sample_quantity=%s", len(out["test_items"]), out["sample_quantity"])
    return out
